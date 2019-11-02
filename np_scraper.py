# Check the permit status of 3 Taiwan National Parks
# --------------------------------------------------------------------------------------------------------------
# Auto querying data from Nation Park Permit Application System.
# --------------------------------------------------------------------------------------------------------------
# Author: David Wang
# Copyright 2019

import argparse
import pandas as pd
from pandas import ExcelWriter
import openpyxl
from openpyxl import load_workbook
import requests
import re
from lxml import html
from bs4 import BeautifulSoup
from selenium import webdriver
from datetime import date, datetime, timedelta
from time import sleep
from socket import error as SocketError
import multiprocessing
from multiprocessing import Pool
from multiprocessing.pool import ThreadPool
import functools
from requests.auth import HTTPProxyAuth

# Set up arguments for this program
parser = argparse.ArgumentParser(description="Have fun dude!", formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument("-p", "--park", required=True, help="<Required> National Park that you plan to visit")
parser.add_argument('-l','--lodge', required=True, nargs='+', help="<Required> One or more Lodge/Campsite where you plan to stay. For example: -l 三六九山莊 七卡山莊")
parser.add_argument("-s", "--start", required=False, help="Use this parameter to query single date status. The query date must be 7 days later. Format example: 2019-02-02")
parser.add_argument("-e", "--end", required=False, help="Use this parameter to query multiple date from startDate to endDate. Format example: 2019-02-22")
parser.add_argument("-n", "--number", required=False, help="Team numbers that you plan to apply for permit. For example: -n 5")
parser.add_argument("-r", "--retain", required=False, help="Check if still have retained room number for foreigner. Follow the -n argument. For example: -n 5 -r yes")

# Combine all arguments into a list called args
args = parser.parse_args()
national_park = args.park
lodge_campsite = args.lodge
start_date = args.start
end_date = args.end
team_number = args.number
check_retain = args.retain

class TaiwanNationalParkWebParser:
    def append_df_to_excel(self, filename, df, summarize, summarize_string, sheet_name='Selection', startrow=None, truncate_sheet=False, **to_excel_kwargs):
        """Save table data to Excel/CSV or PDF file."""
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row + 1

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out table data to the sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # write the summarize data below table data
        newrow = writer.book[sheet_name].max_row

        # define summarize data string and write it to the sheet
        writer.book[sheet_name].cell(row = newrow+1, column = 1).value = summarize_string

        # save the workbook
        writer.save()

    def parse_orgid(self, url):
        """Parsing org_id for specific National Park."""
        headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36'}
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                resp.encoding = resp.apparent_encoding
                soup = BeautifulSoup(resp.text, 'lxml')
                items = soup.select('ul > li > a[href^="apply_1_2.aspx?unit="]')
                for i in items:
                    if national_park not in str(i):
                        continue
                    # Get orgid for selected national park
                    return str(i.get('href').replace('apply_1_2.aspx?unit=', ''))
        except Exception as ex:
            print(str(ex))

    def get_lodge_list(self, url):
        """Get the lodge/campsite list and its id by National Park."""
        headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36'}
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                resp.encoding = resp.apparent_encoding
                soup = BeautifulSoup(resp.text, 'lxml')
                items = soup.select('option[value^=""]')
                lodge_ids = [item.get('value') for item in items]
                lodge_names = [item.text for item in items]
                lodge_camp_list = dict(zip(lodge_names, lodge_ids))
                return lodge_camp_list
        except Exception as ex:
            print(str(ex))

    def parse_url(self, url, np, lodge_id):
        """Parsing web content, and extract the required information from XML and HTML structures."""
        headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36'}
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                resp.encoding = resp.apparent_encoding
                soup = BeautifulSoup(resp.text, 'lxml')
                # First check if this page exist
                if soup.find("span", id="ContentPlaceHolder1_sdate") == None:
                    print("您所查詢的宿營地，於該日未開放查詢！")
                    return {lodge_id:0}
                # Parsing web data
                search_date = soup.find("span", id="ContentPlaceHolder1_sdate").text
                national_park = soup.find("span", id="ContentPlaceHolder1_org").text
                lodge = soup.find("span", id="ContentPlaceHolder1_room").text
                table = soup.find("table", class_="DATAM")
                #lodge_available_list[lodge_id] = []
                if np == "玉山":
                    # 餘額
                    current_available = soup.find("span", id="ContentPlaceHolder1_lbCnt1").text
                    current_available = current_available.replace('(', '').replace(')', '').split(',')[0]
                    # 乘載量
                    pool_total = soup.find("span", id="ContentPlaceHolder1_lbCnt").text
                    pool_total = pool_total.replace('(', '').replace(')', '').split(',')[0]
                    # 排隊預約
                    queue = soup.find("span", id="ContentPlaceHolder1_lbStatus_6").text
                    # 審核中
                    examine = soup.find("span", id="ContentPlaceHolder1_lbCnt2").text
                    # 核准入園
                    approved = soup.find("span", id="ContentPlaceHolder1_lbStatus_4").text
                    # 共計
                    total_applicant = int(queue) + int(examine)
                    # 中籤率
                    if int(queue) == 0 and int(examine) == 0 and int(approved) == 0:
                        # The date has preserved and can't apply permit
                        percentage = 'N/A'
                    else:
                        # The drew lots result not public yet (Date after 30 days)
                        if int(approved) == 0 and int(queue) > 0:
                            percentage = 100*(int(pool_total) / int(queue))
                            percentage = 100 if percentage > 100 else percentage
                        else:
                            if team_number and not check_retain:
                                percentage = 100 if int(pool_total) - (int(queue) + int(examine) + int(approved)) > int(team_number) else 0
                            else:
                                percentage = '已抽完籤'

                    summarize = {'search_date': search_date,'national_park' : national_park, 'lodge' : lodge, 'current_available' : current_available, 'pool_total' : pool_total, 'queue' : queue, 'examine' : examine, 'approved' : approved, 'total_applicant' : total_applicant, 'percentage' : percentage}
                    summarize_string = "{} {} {}\n餘額：{} │ 承載量：{} │ 排隊預約： {}位 │ 審核中： {}位 │ 核准入園：{}位 ，共計：{}位，中籤率約為 {} %。\n".format(search_date, national_park, lodge, current_available, pool_total, queue, examine, approved, total_applicant, percentage)

                    # Check if can apply permit by team number
                    if team_number and not check_retain:
                        if int(current_available) > int(queue) and int(current_available) - int(queue) >= int(team_number):
                            if soup.find("table", class_="DATAM") == None:
                                lodge_available_list.update({lodge_id:0})
                                #lodge_available_list.update({lodge_id : {search_date : 0}})
                            else:
                                lodge_available_list.update({lodge_id:1})
                                #lodge_available_list.update({lodge_id : {search_date : 1}})
                        else:
                            # Date that not draw lots yet. Everyone can still apply permit
                            if int(queue) > 0 and int(examine) == 0 and int(approved) == 0:
                                lodge_available_list.update({lodge_id:1})
                                #lodge_available_list.update({lodge_id : {search_date : 1}})
                            else:
                                lodge_available_list.update({lodge_id:0})
                                #lodge_available_list.update({lodge_id : {search_date : 0}})

                    # Check if can apply retain permit by foreigner team number
                    if team_number and check_retain:
                        if table != None:
                            retain_number = 0
                            for row in table.find_all("tr"):
                                columns = row.find_all('td')
                                combinaiton = dict(enumerate(columns))
                                if "外籍提前保留名額" in str(combinaiton.get(9)) or "外籍提前申請" in str(combinaiton.get(9)):
                                    retain_number += int(combinaiton.get(6).get_text())
                            if 24 - int(retain_number) >=  int(team_number):
                                lodge_available_list.update({lodge_id:1})
                                print("尚餘可申請外籍保留名額：{}位".format(24 - int(retain_number)))
                            else:
                                lodge_available_list.update({lodge_id:0})
                                print("尚餘可申請外籍保留名額：0位")
                        else:
                            lodge_available_list.update({lodge_id:0})
                elif np == "雪霸":
                    # 乘載量
                    pool_total = soup.find("span", id="ContentPlaceHolder1_lblsumrooms").text
                    # 待處理
                    queue = soup.find("span", id="ContentPlaceHolder1_lblchkrooms").text
                    # 補件
                    wait = soup.find("span", id="ContentPlaceHolder1_docpeople").text
                    # 核准入園
                    approved = soup.find("span", id="ContentPlaceHolder1_lblsubrooms").text
                    # 待系統排定
                    tbd = soup.find("span", id="ContentPlaceHolder1_lblsystemwait").text
                    # 候補
                    candidate = soup.find("span", id="ContentPlaceHolder1_lblbakrooms").text
                    # 餘額
                    current_available = soup.find("span", id="ContentPlaceHolder1_lbloverrooms").text
                    summarize = {'search_date': search_date,'national_park' : national_park, 'lodge' : lodge, 'pool_total' : pool_total, 'queue' : queue, 'wait' : wait, 'approved' : approved, 'tbd' : tbd, 'candidate' : candidate, 'current_available' : current_available}
                    summarize_string = "{} {} {}\n餘額：{}床位 │ 承載量：{} │ 待處理： {}床位 │ 補件： {}床位 │ 已通過：{}床位 | 待系統排定：{}床位 │ 宿營地不足後補：{}床位\n".format(search_date, national_park, lodge, current_available, pool_total, queue, wait, approved, tbd, candidate)

                    # Check if you can apply permit with team_number in the preferred date range
                    if team_number:
                        if int(current_available) > 0 and int(current_available) - int(tbd) >= int(team_number):
                            lodge_available_list.update({lodge_id:1})
                        else:
                            lodge_available_list.update({lodge_id:0})
                elif np == "太魯閣":
                    # 乘載量
                    pool_total = soup.find("span", id="ContentPlaceHolder1_lblsumrooms").text
                    # 核准入園
                    approved = soup.find("span", id="ContentPlaceHolder1_lblsubrooms").text
                    # 待審核
                    tbd = soup.find("span", id="ContentPlaceHolder1_lblchkrooms").text
                    # 餘額
                    current_available = soup.find("span", id="ContentPlaceHolder1_lbloverrooms").text
                    summarize = {'search_date': search_date,'national_park' : national_park, 'lodge' : lodge, 'pool_total' : pool_total, 'approved' : approved, 'tbd' : tbd, 'current_available' : current_available}
                    summarize_string = "{} {} {}\n餘額：{}床位 │ 承載量：{} │ 通過審核： {}床位 │ 待審核： {}床位\n".format(search_date, national_park, lodge, current_available, pool_total, approved, tbd)

                    # Check if you can apply permit with team_number in the preferred date range
                    if team_number:
                        if int(current_available) > 0 and int(current_available) >= int(team_number):
                            lodge_available_list.update({lodge_id:1})
                        else:
                            lodge_available_list.update({lodge_id:0})

                #Parsing and return detail table data from webpage
                if table != None:
                    print(summarize_string, [(self.parse_html_table(soup.find_all("table", class_="DATAM")[0], summarize, summarize_string))])
                else:
                    print(summarize_string, "已保留{}{}床位，供活動暨工作人員使用".format(search_date, lodge))

                return lodge_available_list

        except Exception as ex:
            print(str(ex))

    def parse_html_table(self, table, summarize, sum_string):
        """Parse table data from web site."""
        n_columns = 0
        n_rows=0
        column_names = []

        # Find number of rows and columns
        # we also find the column titles if we can
        for row in table.find_all('tr'):
            # Determine the number of rows in the table
            td_tags = row.find_all('td')
            if len(td_tags) > 0:
                n_rows+=1
                if n_columns == 0:
                    # Set the number of columns for our table
                    n_columns = len(td_tags)

            # Handle column titles if we find them
            title_tags = row.find_all("th")
            if len(title_tags) > 0: #and len(column_names)-1 == 0:
                for th in title_tags:
                    column_names.append(th.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))

        # Safeguard on Column Titles
        if len(column_names) > 0 and len(column_names) != n_columns:
            #raise Exception("Column titles do not match the number of columns")
            print("此日期尚無隊伍資料，或尚未開放申請入園")

        columns = column_names if len(column_names) > 0 else range(0, n_columns)
        df = pd.DataFrame(columns = columns, index= range(0, n_rows))
        row_marker = 0
        #preserved = 0
        for row in table.find_all("tr"):
            column_marker = 0
            columns = row.find_all('td')
            #combinaiton = dict(zip(column_names, columns))
            #print("combinaiton =", combinaiton)
            #if "外籍提前保留名額" in str(combinaiton.get(column_names[9])):
            #    preserved += int(combinaiton.get(column_names[6]).get_text())

            for column in columns:
                df.iat[row_marker, column_marker] = column.get_text()
                column_marker += 1
            if len(columns) > 0:
                row_marker += 1
        #print("preserved=", preserved)

        # Write Pandas DF to Excel
        excel_file = "山屋申請查詢.xlsx"
        sheet_name = "入園申請查詢表"
        #self.append_df_to_excel(excel_file, df, summarize, sum_string, sheet_name=sheet_name)

        #writer = ExcelWriter('WebDataExport.xlsx')
        #df.to_excel(writer, 'SharesSelection')
        #writer.save()

        # Write Pandas DF to CSV
        #df.to_csv('WebDataExport.csv', sep=',')

        return df

    def check_available_apply_date(self):
        """Check available apply date by given team number, lodge/campsite arrangement and date range."""
        available_date_list = []
        available_date = []
        for item in lodge_available_list.values():
            available_date_list.append(item)

        N = 1 if len(date_range)-len(lodge_campsite) == 0 else len(date_range)-len(lodge_campsite)
        for n in range(N):
            tmp = []
            [tmp.append(available_date_list[a][n+a]) for a in range(len(available_date_list))]
            #for a in range(len(available_date_list)):
            #    tmp.append(available_date_list[a][n+a])
            if tmp == [1] * len(lodge_campsite):
                # Filtering only that are series days between Sunday and Thrusday and have retained numbers
                if check_retain and len(lodge_campsite) > 1:
                    tmp = date_range[n+len(lodge_campsite)].split('-')
                    tmp[0] = str(int(tmp[0])+1911)
                    common_date_end = "-".join(tmp)

                    tmp = date_range[n].split('-')
                    tmp[0] = str(int(tmp[0])+1911)
                    common_date_start = "-".join(tmp)

                    day_range = datetime.strptime(common_date_end, '%Y-%m-%d') - datetime.strptime(common_date_start, '%Y-%m-%d')
                    if  day_range.days > len(lodge_campsite):
                        continue
                available_date.append(date_range[n])
                print(date_range[n], "尚可申請入園")
            else:
                tmp.clear()
        return available_date


# In the first instance we check the web site connection
check_url = 'https://npm.cpami.gov.tw/bed_menu.aspx'
headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36'}
try:
    r = requests.get(check_url, timeout=10)
    if r.status_code == 200:
        print('Check connect OK')
except Exception as ex:
    print(str(ex))
    exit(0)

# Check if start_data and end_date correctly
if check_retain:
    check_start_date = datetime.strftime(date.today() + timedelta(days=35), "%Y-%m-%d")
    check_end_date = datetime.strftime(date.today() + timedelta(days=120), "%Y-%m-%d")
    if  start_date and start_date < check_start_date or end_date and end_date > check_end_date:
        print("外籍保留名額提前申請期限：預定入園日期前35天至前4個月間提出申請")
        print("外籍保留名額可申請日期為：{} 到 {}".format(check_start_date, check_end_date))
        exit(0)

# Set start_data and end_date
if start_date and end_date:
    day_range = datetime.strptime(end_date, '%Y-%m-%d') - datetime.strptime(start_date, '%Y-%m-%d')
    if day_range.days < len(lodge_campsite):
        print("您輸入了{}個山屋/營地名稱，請至少以{}天行程來查詢床位".format(len(lodge_campsite), len(lodge_campsite)+1))
        exit(0)
elif not start_date and not end_date:
    if check_retain:
        start_date = check_start_date
        end_date = check_end_date
    else:
        start_date = datetime.strftime(date.today() + timedelta(days=7), "%Y-%m-%d")
        end_date = datetime.strftime(date.today() + timedelta(days=28), "%Y-%m-%d")
    print("你沒有輸入入園和下山日期，將自動搜尋{}到{}之間的入園申請狀況".format(start_date, end_date))
elif start_date and not end_date:
    if team_number:
        end_date = datetime.strftime(datetime.strptime(start_date, '%Y-%m-%d') + timedelta(days=len(lodge_campsite)), "%Y-%m-%d")
        print("你沒有輸入下山日期，將自動搜尋{}到{}之間的入園申請狀況".format(start_date, end_date))
else:
    print("您輸入了{}個山屋/營地名稱，請選擇入園日期、下山日期、隊伍人數，來查詢可申請時段\n或者不輸入入園日期、下山日期，將會自動搜尋7日以後到30日以內的所有可申請時段".format(len(lodge_campsite)))
    exit(0)

# Define National Park name and Lodge lists
national_park_list = ["玉山", "太魯閣", "雪霸"]
for np_item in national_park_list:
    national_park = national_park.replace("國家公園", "")
    if national_park in national_park_list:
        if '玉山' == national_park:
            get_lodge_link = "bed_6.aspx"
            check_bed_link = "bed_6main.aspx"
        elif '雪霸' in national_park:
            get_lodge_link = "bed_1.aspx"
            check_bed_link = "bed_1main.aspx"
        elif '太魯閣' in national_park:
            get_lodge_link = "bed_4.aspx"
            check_bed_link = "bed_4main.aspx"

        # Check if lodge name correct
        hp = TaiwanNationalParkWebParser()
        check_lodge_list_link = "https://npm.cpami.gov.tw/{}".format(get_lodge_link)
        lodge_camp_list = hp.get_lodge_list(check_lodge_list_link)
        for lodge in lodge_campsite:
            if lodge not in lodge_camp_list:
                print("{} <--名稱錯誤。\n請輸入正確的山屋/營地名稱。{}國家公園路線的山屋/營地如下：\n{}".format(lodge, national_park, list(lodge_camp_list.keys())))
                exit(0)
    else:
        print("請輸入正確的國家公園名稱，例如：玉山 / 太魯閣 / 雪霸")
        exit(0)

# Check if input date format correctly
try:
    datetime.strptime(start_date, '%Y-%m-%d')
    if end_date:
        datetime.strptime(end_date, '%Y-%m-%d')
except ValueError:
    raise ValueError("Incorrect data format, should be YYYY-MM-DD")

available_start_date = datetime.strftime(date.today() + timedelta(days=7), "%Y-%m-%d")
if start_date >= available_start_date:
    date_range = []
    # If endDate parameter provided, get the date list between startDate and endDate
    if end_date:
        if end_date > start_date:
            delta = datetime.strptime(end_date, "%Y-%m-%d").date() - datetime.strptime(start_date, "%Y-%m-%d").date()
            for i in range(delta.days + 1):
                #print(datetime.strptime(start_date, "%Y-%m-%d").date() + timedelta(i))
                the_day = str(datetime.strptime(start_date, "%Y-%m-%d").date() + timedelta(i))
                tmp = the_day.split("-")
                dayOfWeek = datetime.strptime(the_day, "%Y-%m-%d").weekday()
                # Ignore checking 'Friday' and 'Saturday' for querying foreigner retain number
                if team_number and check_retain and (dayOfWeek == 4 or dayOfWeek == 5):
                    continue
                tmp[0] = str(int(tmp[0])-1911)
                chinese_date = "-".join(tmp)
                date_range.append(chinese_date)
        else:
            print("欲查詢的結束日期必須在開始日期之後")
            exit(0)
    else:
        tmp = start_date.split("-")
        tmp[0] = str(int(tmp[0])-1911)
        chinese_date = "-".join(tmp)
        date_range.append(chinese_date)
else:
    print("僅提供7日以後申請案之進度查詢")
    exit(0)


lodge_available_list = {}
check_orgid_url = "https://npm.cpami.gov.tw/bed_menu.spx"
hp = TaiwanNationalParkWebParser()
orgid = hp.parse_orgid(check_orgid_url)

# Set process count
if len(date_range) > multiprocessing.cpu_count():
    pc = len(date_range)
else:
    pc = multiprocessing.cpu_count()

# Implement multiprocessing Pool to improve the performance
for lodge in range(len(lodge_campsite)):
    checkList = []
    inquire_url_list = []
    for n in date_range:
        inquire_url = "https://npm.cpami.gov.tw/{}?orgid={}&node_id={}&sdate={}".format(check_bed_link, orgid, lodge_camp_list[lodge_campsite[lodge]], n)
        inquire_url_list.append(inquire_url)
    #pool = Pool(processes=multiprocessing.cpu_count())
    pool = Pool(processes=multiprocessing.cpu_count())
    p = pool.map(functools.partial(hp.parse_url, np=national_park, lodge_id=lodge), inquire_url_list)
    pool.terminate()
    pool.join()
    [checkList.append(i.get(lodge)) for i in p]
    lodge_available_list.update({lodge : checkList})

if team_number:
    available_date = hp.check_available_apply_date()
    if len(available_date) > 0:
        print("隊伍{}人，共{}個時段可申請入園".format(team_number, len(available_date)))
    else:
        print("沒有任何可申請入園的時段")
