import pickle
from datetime import date, datetime, timedelta
import os
import re
from urllib.parse import urlparse
import requests
from selenium import webdriver
from lxml import html
from bs4 import BeautifulSoup
# Save data to Pandas DataFrame and write to Excel/CSV file
import pandas as pd
from pandas import ExcelWriter
import openpyxl
from openpyxl import load_workbook
# Multiprocessing
import multiprocessing
from multiprocessing import Pool
from multiprocessing.pool import ThreadPool
import functools

class LodgeRoomChecker:
    def __init__(self,
                 userAgent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36',
                 **kwargs):
        self.userAgent = userAgent

    def parse_csrf(self, url):
        """Parsing CSRF code from web page."""
        headers = {'user-agent': self.userAgent}
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                resp.encoding = resp.apparent_encoding
                soup = BeautifulSoup(resp.text, 'lxml')
                csrf = soup.find('form', {'name': 'form1'}).find('input', {'name': 'csrf'}).get('value')
                return csrf
        except Exception as ex:
            print(str(ex))

    def parse_url(self, check_date, lodge_id, lodge_list, team_number, url, csrf):
        """Parsing web content, and extract the required information from XML and HTML structures."""
        check_year = check_date.split('-')[0]
        check_month = check_date.split('-')[1]
        url += '?date_set%5Byear%5D={}&date_set%5Bmonth%5D={}&csrf={}#main2'.format(check_year, check_month, csrf)
        headers = {'user-agent': self.userAgent}
        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code == 200:
            resp.encoding = resp.apparent_encoding
            soup = BeautifulSoup(resp.text, 'lxml')
            calendar_table = soup.find("table", {"class":"calendar_table"}).find("table").find_all("table")
            data_dict = {}
            lodge_available_list = {}
            for row in calendar_table:
                data = []
                day_num = row.find_all("tr")[0].get_text().strip()
                info = row.find_all("tr")[1].find_all("font")
                # Parsing contents in each single day
                for i in info:
                    if i.next_sibling == None:
                        data.append((i.get_text().strip(), ''))
                    else:
                        numberStr = i.next_sibling.string.replace(' ', '').replace(':', '').replace('(', '').replace(')', '').strip()
                        data.append((i.get_text().strip(), numberStr))
                data_dict.update({day_num:data})

            # Print out the information
            dayNum = str(int(check_date.split('-')[2]))
            day_info_set = data_dict.get(dayNum)
            if len(day_info_set) == 0:
                summarize = "{} │ 所有床位/營地已額滿".format(check_date)
            else:
                summarize = "{}".format(check_date)
                for i in day_info_set:
                    summarize += " │ {} {}".format(i[0], i[1])

            print(summarize)

            if team_number:
                d = str(datetime.strptime(check_date, '%Y-%m-%d').day)
                if len(data_dict.get(d)) == 0 or len(data_dict.get(d)) == 1:
                    # Room bed is full
                    current_available = 0
                else:
                    # Get available room bed number
                    lodge_name = lodge_list[lodge_id]
                    idx = data_dict.get(d).index((lodge_name, ''))
                    current_available = data_dict.get(d)[idx + 1][1]
                    #current_available = data_dict.get(d)[1][1]
                # Mark room bed of this date is available or not
                if int(current_available) >= int(team_number):
                    lodge_available_list.update({lodge_id:1})
                else:
                    lodge_available_list.update({lodge_id:0})

        return lodge_available_list

    def check_available_apply_date(self, lodge_available_list={}, lodge_campsite=[]):
        """Check available apply date by given team number, lodge/campsite arrangement and date range."""
        print('lodge_available_list:', lodge_available_list)
        available_date_list = []
        available_date = []
        for item in lodge_available_list.values():
            available_date_list.append(item)

        N = 1 if len(date_range)-len(lodge_campsite) == 0 else len(date_range)-len(lodge_campsite)
        for n in range(N):
            tmp = []
            [tmp.append(available_date_list[a][n+a]) for a in range(len(available_date_list))]
            #for a in range(len(available_date_list)):
            #    tmp.append(available_date_list[a][n+a])
            if tmp == [1] * len(lodge_campsite):
                available_date.append(date_range[n])
                print(date_range[n], "可申請入園")
            else:
                tmp.clear()
        return available_date

class MyLoginSession:
    """
    a class which handles and saves login sessions. It also keeps track of proxy settings.
    It does also maintine a cache-file for restoring session data from earlier
    script executions.
    """
    def __init__(self,
                 loginUrl,
                 loginData,
                 loginTestUrl,
                 loginTestString,
                 sessionFileAppendix = '_session.dat',
                 maxSessionTimeSeconds = 1 * 60,
                 proxies = None,
                 userAgent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36',
                 debug = True,
                 forceLogin = False,
                 **kwargs):
        """
        save some information needed to login the session

        you'll have to provide 'loginTestString' which will be looked for in the
        responses html to make sure, you've properly been logged in

        'proxies' is of format { 'https' : 'https://user:pass@server:port', 'http' : ...
        'loginData' will be sent as post data (dictionary of id : value).
        'maxSessionTimeSeconds' will be used to determine when to re-login.
        """
        urlData = urlparse(loginUrl)

        self.proxies = proxies
        self.loginData = loginData
        self.loginUrl = loginUrl
        self.loginTestUrl = loginTestUrl
        self.maxSessionTime = maxSessionTimeSeconds
        self.sessionFile = urlData.netloc + sessionFileAppendix
        self.userAgent = userAgent
        self.loginTestString = loginTestString
        self.debug = debug

        self.login(forceLogin, **kwargs)

    def modification_date(self, filename):
        """
        return last file modification date as datetime object
        """
        t = os.path.getmtime(filename)
        return datetime.fromtimestamp(t)

    def login(self, forceLogin = False, **kwargs):
        """
        login to a session. Try to read last saved session from cache file. If this fails
        do proper login. If the last cache access was too old, also perform a proper login.
        Always updates session cache file.
        """
        wasReadFromCache = False
        if self.debug:
            print('loading or generating session...')
        if os.path.exists(self.sessionFile) and not forceLogin:
            time = self.modification_date(self.sessionFile)

            # only load if file less than 30 minutes old
            lastModification = (datetime.now() - time).seconds
            if lastModification < self.maxSessionTime:
                with open(self.sessionFile, "rb") as f:
                    self.session = pickle.load(f)
                    wasReadFromCache = True
                    if self.debug:
                        print("loaded session from cache (last access %ds ago) "
                              % lastModification)
        if not wasReadFromCache:
            self.session = requests.Session()
            self.session.headers.update({'user-agent' : self.userAgent})
            res = self.session.post(self.loginUrl, data = self.loginData,
                                    proxies = self.proxies, **kwargs)
            if self.debug:
                print('created new session with login' )
            self.saveSessionToCache()

        # test login
        res = self.session.get(self.loginTestUrl)
        if res.text.lower().find(self.loginTestString.lower()) < 0:
            raise Exception("could not log into provided site '%s'"
                            " (did not find successful login string)"
                            % self.loginUrl)

    def saveSessionToCache(self):
        """
        save session to a cache file
        """
        # always save (to update timeout)
        with open(self.sessionFile, "wb") as f:
            pickle.dump(self.session, f)
            #if self.debug:
                #print('updated session cache-file %s' % self.sessionFile)

    def retrieveContent(self, url, method = "get", postData = None, **kwargs):
        """
        return the content of the url with respect to the session.

        If 'method' is not 'get', the url will be called with 'postData'
        as a post request.
        """
        if method == 'get':
            res = self.session.get(url , proxies = self.proxies, **kwargs)
        else:
            res = self.session.post(url , data = postData, proxies = self.proxies, **kwargs)

        # the session has been updated on the server, so also update in cache
        self.saveSessionToCache()

        return res

    def parse_order_detail_1(self, table):
        """Parse order detail and get member lists"""
        column_names = []
        tmp_values = []
        for row in table.find_all('tr'):
            td_columns = row.find_all('td')
            # Handle column titles if we find them
            title_tags = row.find_all("th")
            if len(title_tags) > 0:
                for th in title_tags:
                    column_names.append(th.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))
            if len(td_columns) > 0:
                for td in td_columns:
                    columns = column_names if len(column_names) > 0 else range(0, len(column_names))
                    tmp_values.append(td.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))

        df = pd.DataFrame(columns = columns, index= range(0, 1))

        row_marker = 0
        for row in table.find_all("tr"):
            column_marker = 0
            columns = row.find_all('td')
            for column in columns:
                df.iat[column_marker, row_marker] = column.get_text()
                column_marker += 1
            if len(columns) > 0:
                row_marker += 1

        # Write Pandas DF to Excel
        excel_file = "床位訂單查詢.xlsx"
        sheet_name = "床位訂單查詢"
        #self.append_df_to_excel(excel_file, df, sheet_name=sheet_name)

        return df

    def parse_order_detail_2(self, table):
        """Parse order detail and get member lists"""
        column_names = []
        member_list = []
        n_columns = 0
        for idx, row in enumerate(table.find_all('tr')):
            if idx is 2:
                # Handle column titles if we find them
                title_tags = row.find_all("th")
                if len(title_tags) > 0:
                    for th in title_tags:
                        column_names.append(th.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))

        for idx, row in enumerate(table.find_all('tr')):
            tmp = []
            if idx == 0 or idx == 1:
                member_num = row.find_all("th")
                td_columns = row.find_all('td')
                for th in member_num:
                    tmp.append(th.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))
                for td in td_columns:
                    tmp.append(td.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))
                n_columns = len(tmp)
                member_list.append(tmp)
            elif idx > 2:
                member_num = row.find_all("th")
                td_columns = row.find_all('td')
                if len(member_num) + len(td_columns)  == len(column_names):
                    for th in member_num:
                        tmp.append(th.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))
                    for td in td_columns:
                        tmp.append(td.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))
                    n_columns = len(tmp)
                    member_list.append(tmp)

        columns = column_names if len(column_names) > 0 else range(0, len(n_columns))
        df = pd.DataFrame(columns = columns, index= range(0, len(member_list)))

        row_marker = 0
        for single_member in member_list:
            column_marker = 0
            for member_item in single_member:
                df.iat[row_marker, column_marker] = member_item
                column_marker += 1
            if len(columns) > 0:
                row_marker += 1

        # Write Pandas DF to Excel
        excel_file = "床位訂單查詢.xlsx"
        sheet_name = "床位訂單查詢"
        #self.append_df_to_excel(excel_file, df, sheet_name=sheet_name)

        return df

    def parse_order_table(self, table):
        """Parse order lists from web site."""
        column_names = []
        tmp = []

        for row in table.find_all('tr'):
            td_columns = row.find_all('td')
            # Handle column titles if we find them
            title_tags = row.find_all("th")
            if len(title_tags) > 0:
                for th in title_tags:
                    column_names.append(th.get_text().replace('\r', '').replace('\n', '').replace('\t', ''))

            if len(td_columns) == len(column_names):
                tmp.append(td_columns)

            columns = column_names if len(column_names) > 0 else range(0, len(column_names))
            df = pd.DataFrame(columns = columns, index= range(0, len(tmp)))

        row_marker = 0
        td_sets_list = {}
        for order_item in tmp:
            column_marker = 0
            can_be_delete = False
            for td in order_item:
                # Get the SID, if this order could be modify
                items = td.find_all(lambda tag: tag.name == "a" and (tag.has_attr("onclick") and tag.has_attr("href")))
                for i in items:
                    if re.match('checkDel', i['onclick']):
                        order_sid = re.findall("(\\w+)", i['onclick'])[1]
                        can_be_delete = True
                if can_be_delete:
                    df.iat[row_marker, column_marker] = order_sid
                else:
                    df.iat[row_marker, column_marker] = td.get_text().replace('\r', '').replace('\n', '').replace('\t', '')
                column_marker += 1

            if len(order_item) > 0:
                row_marker += 1

            td_item_sets = [td.get_text().replace('\r', '').replace('\n', '').replace('\t', '') for td in order_item]
            td_sets_list.update({td_item_sets[2]:td_item_sets})

        # Write Pandas DF to Excel
        excel_file = "我的床位訂單.xlsx"
        sheet_name = "床位訂單查詢"
        #self.append_df_to_excel(excel_file, df, sheet_name=sheet_name)

        return df, td_sets_list

    def append_df_to_excel(self, filename, df, sheet_name='Selection', startrow=None, truncate_sheet=False, **to_excel_kwargs):
        """Save table data to Excel/CSV or PDF file."""
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row + 1

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out table data to the sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()


def parse_information(postData):
    resp = s.retrieveContent('https://npm.cpami.gov.tw/apply_2.aspx', method = "post", postData = postData)
    resp.encoding = resp.apparent_encoding
    #print(postData.get('ctl00$ContentPlaceHolder1$serial'))
    if resp.status_code == 200 and not "查無資料" in resp.text:
        appNum = postData.get('ctl00$ContentPlaceHolder1$serial')
        print('入園申請編號：' + appNum)

if __name__ == "__main__":
    # Define data payload
    loginData = {'is_uu':'登入帳號', 'is_pp':'登入密碼', 'mode':'log_in'}
    applyData = {'mode':'insert', 'this_mode':'add', 'user_id':'', 'member_num':'1', 'room_type':'2', 'room_type_2':'0', 'name':'姓名', 'phone_day':'電話', 'phone_night':'', 'mobile':'手機號碼', 'zone':'郵遞區號', 'address':'地址', 'email':'電郵地址', 'total_order_day':'2', 'room_subid_1[]':'13', 'room_subid_2[]':'11', 'check_num_1[13]':'1', 'check_num_2[11]':'1', 'have_alpine':'0', 'roomqty_1[]':'1', 'roomqty_2[]':'1', 'num_2[11]':'1', 'num[13]':'1', 'team_detail[0][leader]':'1', 'team_name':'隊伍名稱', 'date':'2019-08-05', 'date_start':'2019-08-05', 'date_end':'2019-08-07', 'date_start_2':'2019-08-05', 'date_end_2':'2019-08-07', 'team_detail[0][name]':'姓名', 'team_detail[0][idnumber]':'身分證', 'team_detail[0][birth]':'生日', 'team_detail[0][e_name]':'姓名', 'team_detail[0][e_tel]':'手機號碼', 'team_detail[0][note]':'', 'payment':'24', 'payment_default':'1', 'date_end_ok':'2019-08-06', 'lnvoice_DonateMark':'0', 'schedule':'D1:登山口~向陽山屋\nD2:向陽山屋~嘉明湖~避難山屋\nD3:避難山屋~登山口~下山回家', 'check_here':'ok', 'Submit':'送出申請單'}

    # Define string and url
    loginUrl = 'https://jmlnt.forest.gov.tw/members/?mode=sign_in'
    loginTestUrl = 'https://jmlnt.forest.gov.tw/members/index.php'
    successStr = '歡迎來到會員專區'
    applyCompleteStr = '申請單成功 , 請等待抽籤 !'
    duplicateAlertStr = '請勿重複訂同一天'
    orderCanceledStr = '訂單已取消'
    orderPayedStr = '已收款'
    errOperationStr = '操作錯誤'

    applyUrl = "https://jmlnt.forest.gov.tw/room/index.php?type=roomsubmit#focus"
    orderUrl = "https://jmlnt.forest.gov.tw/room/index.php?mode=record"
    deleteOrderUrl = "https://jmlnt.forest.gov.tw/room/index.php?mode=del&id=xxxxx"

    # Input parameter
    team_number = 3
    lodge_campsite = ['檜谷山莊']
    date_range = ['2019-06-14']

    # Parsing and get CSRF code from web page
    check_room_url = "https://jmlnt.forest.gov.tw/room/index.php"
    hp = LodgeRoomChecker()
    csrf = hp.parse_csrf(check_room_url)

    # lodge_available_list = {}
    # for lodge_id in range(len(lodge_campsite)):
    #     checkList = []
    #     pool = Pool(processes=10)
    #     p = pool.map(functools.partial(hp.parse_url, lodge_id=lodge_id, lodge_list=lodge_campsite, team_number=team_number, url=check_room_url, csrf=csrf), date_range)
    #     pool.terminate()
    #     pool.join()
    #     [checkList.append(i.get(lodge_id)) for i in p]
    #     lodge_available_list.update({lodge_id : checkList})

    # if team_number:
    #     available_date = hp.check_available_apply_date(lodge_available_list=lodge_available_list, lodge_campsite=lodge_campsite)
    # if len(available_date) > 0:
    #     print("隊伍{}人，共{}個時段可申請入園".format(team_number, len(available_date)))
    # else:
    #     print("沒有任何可申請入園的時段")



    s = MyLoginSession(loginUrl, loginData, loginTestUrl, successStr,
                       #proxies = proxies
                       )

    # Get the member list detail from specific order
    # for x in range(108006005, 108099999):
    #     num = "Y{}".format(x)
    #     print(num)
    #     tryApplyData = {'ctl00$ContentPlaceHolder1$serial':num, 'ctl00$ContentPlaceHolder1$nation':'中華民國', 'ctl00$ContentPlaceHolder1$sid':'身分證字號', 'ctl00$ContentPlaceHolder1$btnok':'確定', '__EVENTTARGET':'', '__EVENTARGUMENT':'', '__LASTFOCUS':'', '__VIEWSTATE':'/wEPDwUJNTQ2NjMwNjAxD2QWAmYPZBYCAgEPZBYGZg8PFgIeC05hdmlnYXRlVXJsBShodHRwczovL25wbS5jcGFtaS5nb3YudHcvZW4vYXBwbHlfMi5hc3B4ZGQCAQ8PFgIfAAUoaHR0cHM6Ly9ucG0uY3BhbWkuZ292LnR3L2pwL2FwcGx5XzIuYXNweGRkAgUPZBYCAgEPZBYCZg9kFgYCBQ8QZA8WA2YCAQICFgMQBQnoq4vpgbjmk4dlZxAFDOS4reiPr+awkeWciwUM5Lit6I+v5rCR5ZyLZxAFBuWci+WklgUG5ZyL5aSWZxYBZmQCCQ8QZGQWAWZkAhMPD2QWAh4Kb25rZXlwcmVzcwULYnRub2tfQ2xpY2tkZNJroTYsMfLdppqB4W+9E6L9a8BNclL9ERuUqZPBN7p7', '__EVENTVALIDATION':'/wEdAAkDTXCCT1Nj6EYNzg6PQ2xkLln1K8JFzMHYcg/+iKQuUTnMTuM25Up1NorzRN5IvuBvGPJbffQ4bZx6UJFVLrb8YzE3nbR8A3zUzMeWMX5hROjoWL/m8Gda+gWz1slMTJQNJfmNJa5ndUmR/4Wu5mq/4l1hz6v11eAuW0dM3caBENsQn0Mp1TXPZtsjU89hOw29ElJmx36YfQIKEtjMDPbEb1BSDCtpzgjKYQ7HyHK4Yw==', '__VIEWSTATEGENERATOR':'C45DDE53'}
    #     resp = s.retrieveContent('https://npm.cpami.gov.tw/apply_2.aspx', method = "post", postData = tryApplyData)
    #     resp.encoding = resp.apparent_encoding

    #     if resp.status_code == 200 and not "查無資料" in resp.text:
    #         print(resp.text)
    #         print('入園申請編號：' + num)


    # Delete an order by order identify number
    # try:
    #     order_num = '1811250008'
    #     resp = s.retrieveContent(orderUrl)
    #     resp.encoding = resp.apparent_encoding
    #     soup = BeautifulSoup(resp.text, 'lxml')
    #     table = soup.find("table", class_="list_table")
    #     order, tsl = s.parse_order_table(table)
    #     if order_num in tsl.keys():
    #         idx = order.index[order['申請單編號'] == order_num][0]
    #         idenfy_num = order['取消'][idx]
    #         if len(idenfy_num) > 0:
    #             # define data payload, and send HTTP request
    #             deleteData = {'id':idenfy_num, 'mode':'del'}
    #             resp = s.retrieveContent(deleteOrderUrl, method = "delete", postData = deleteData)
    #             if resp.status_code == 200:
    #                 resp.encoding = resp.apparent_encoding
    #                 soup = BeautifulSoup(resp.text, 'lxml')
    #                 table = soup.find("table", class_="list_table")
    #                 order, tsl = s.parse_order_table(table)
    #                 if order['申請單狀態'][idx] == orderCanceledStr:
    #                     print(orderCanceledStr)
    #                 elif order['申請單狀態'][idx] == orderPayedStr:
    #                     print(orderPayedStr)
    #                 else:
    #                     print('Something went wrong..')
    #         else:
    #             print('此訂單先前已經被取消了')
    #     else:
    #         print('此訂單編號不存在')
    # except requests.exceptions.RequestException as error:
    #     print(error)


    # Make an new order for room bed
    # try:
    #     res = s.retrieveContent(applyUrl, method = "post", postData = applyData)
    #     if res.text.lower().find(applyCompleteStr.lower()) > 0:
    #         # get order id for further using
    #         resp = s.retrieveContent(orderUrl)
    #         resp.encoding = resp.apparent_encoding
    #         soup = BeautifulSoup(resp.text, 'lxml')
    #         table = soup.find("table", class_="list_table")
    #         order, tsl = s.parse_order_table(table)
    #         print(order)
    #         print(applyCompleteStr)
    #         print('申請單編號:', order['申請單編號'][0])
    #         print('識別碼:', order['取消'][0])
    #     elif res.text.lower().find(duplicateAlertStr.lower()) > 0:
    #         print(duplicateAlertStr)
    #         exit(0)
    # except requests.exceptions.RequestException as error:
    #     print(error)
